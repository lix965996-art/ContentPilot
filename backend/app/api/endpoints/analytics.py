import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_roles
from app.core.exceptions import AppException
from app.core.responses import success_response
from app.db.session import get_db
from app.models.business import ContentArticle, EngagementMetric, PublishSchedule
from app.models.user import User
from app.schemas.business import MetricCreate
from app.services.analytics_service import grouped, summarize
from app.services.audit_service import record_audit

router = APIRouter(tags=["数据复盘"])
HEADERS = [
    "schedule_id",
    "platform",
    "metric_date",
    "group_type",
    "impressions",
    "likes",
    "comments",
    "collects",
    "shares",
    "followers",
    "data_source",
]


def _save_metric(db: Session, payload: MetricCreate) -> EngagementMetric:
    if not db.get(PublishSchedule, payload.schedule_id):
        raise AppException(40407, f"排期任务 {payload.schedule_id} 不存在", 404)
    total = payload.likes + payload.comments + payload.collects + payload.shares
    denominator = payload.impressions or payload.followers
    return EngagementMetric(
        **payload.model_dump(),
        engagement_total=total,
        engagement_rate=round(total / denominator, 6) if denominator else 0,
    )


@router.get("/analytics/template")
def template(_: User = Depends(get_current_user)) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "互动数据导入"
    sheet.append(HEADERS)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=contentpilot-analytics-template.xlsx"
        },
    )


@router.post("/analytics/manual")
def manual(
    payload: MetricCreate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("ADMIN", "OPERATOR")),
) -> dict:
    row = _save_metric(db, payload)
    db.add(row)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise AppException(40911, "同一排期、日期和来源的数据已存在", 409) from exc
    record_audit(
        db, request, user, "CREATE", "ANALYTICS", "METRIC", row.id, {"source": row.data_source}
    )
    db.commit()
    return success_response(
        request,
        {
            "id": row.id,
            "engagementTotal": row.engagement_total,
            "engagementRate": row.engagement_rate,
        },
        "互动数据已保存",
    )


@router.post("/analytics/import")
async def import_metrics(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("ADMIN", "OPERATOR")),
) -> dict:
    if not file.filename or not file.filename.lower().endswith((".csv", ".xlsx")):
        raise AppException(40041, "只支持 CSV 或 XLSX 文件")
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise AppException(40042, "文件不能超过 5MB")
    records: list[dict] = []
    if file.filename.lower().endswith(".csv"):
        try:
            records = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
        except UnicodeDecodeError as exc:
            raise AppException(40043, "CSV 必须使用 UTF-8 编码") from exc
    else:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(x or "") for x in rows[0]] if rows else []
        records = [dict(zip(headers, values, strict=False)) for values in rows[1:]]
    missing = [name for name in HEADERS if name not in (records[0].keys() if records else [])]
    if missing:
        raise AppException(40044, "导入字段不完整", 422, {"missingFields": missing})
    success, errors = 0, []
    for index, item in enumerate(records, start=2):
        try:
            payload = MetricCreate(
                schedule_id=int(item["schedule_id"]),
                platform=str(item["platform"]),
                metric_date=date.fromisoformat(str(item["metric_date"])[:10]),
                group_type=str(item["group_type"]),
                impressions=int(item["impressions"] or 0),
                likes=int(item["likes"] or 0),
                comments=int(item["comments"] or 0),
                collects=int(item["collects"] or 0),
                shares=int(item["shares"] or 0),
                followers=int(item["followers"] or 0),
                data_source=str(item["data_source"] or "IMPORTED"),
            )
            with db.begin_nested():
                db.add(_save_metric(db, payload))
                db.flush()
            success += 1
        except Exception as exc:
            errors.append({"row": index, "message": str(exc)[:200]})
    record_audit(
        db,
        request,
        user,
        "IMPORT",
        "ANALYTICS",
        detail={"filename": file.filename, "success": success, "errors": len(errors)},
    )
    db.commit()
    return success_response(
        request,
        {"successCount": success, "errorCount": len(errors), "errors": errors},
        "数据导入完成",
    )


@router.get("/analytics/overview")
def overview(
    request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)
) -> dict:
    return success_response(request, summarize(db))


@router.get("/analytics/platform-comparison")
def platform_comparison(
    request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)
) -> dict:
    return success_response(request, grouped(db, "platform"))


@router.get("/analytics/time-comparison")
def time_comparison(
    request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)
) -> dict:
    return success_response(request, grouped(db, "group_type"))


@router.get("/analytics/content-ranking")
def content_ranking(
    request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)
) -> dict:
    metrics = db.scalars(
        select(EngagementMetric).order_by(EngagementMetric.engagement_rate.desc()).limit(10)
    ).all()
    items = []
    for metric in metrics:
        schedule = db.get(PublishSchedule, metric.schedule_id)
        article = db.get(ContentArticle, schedule.article_id) if schedule else None
        items.append(
            {
                "scheduleId": metric.schedule_id,
                "title": article.title if article else "未知内容",
                "platform": metric.platform,
                "engagementRate": round(metric.engagement_rate * 100, 2),
                "engagementTotal": metric.engagement_total,
                "dataSource": metric.data_source,
            }
        )
    return success_response(request, items)


@router.get("/analytics/report")
def report(_: User = Depends(get_current_user), db: Session = Depends(get_db)) -> StreamingResponse:
    summary = summarize(db)
    platforms = grouped(db, "platform")
    times = grouped(db, "group_type")
    platform_rows = "".join(
        f"<tr><td>{item['name']}</td><td>{item['sampleCount']}</td>"
        f"<td>{item['engagementRate']}%</td></tr>"
        for item in platforms
    )
    html = (
        "<!doctype html><html lang='zh-CN'><meta charset='utf-8'>"
        "<title>ContentPilot 数据复盘</title><style>body{font:16px/1.7 sans-serif;"
        "max-width:900px;margin:40px auto;color:#172033}table{border-collapse:collapse;"
        "width:100%}td,th{border:1px solid #ddd;padding:8px}</style>"
        f"<h1>ContentPilot 数据复盘报告</h1><p>生成时间：{datetime.now():%Y-%m-%d %H:%M}</p>"
        f"<p>样本数：{summary['sampleCount']}；互动率：{summary['engagementRate']}%；"
        f"其中 SIMULATED：{summary['simulatedCount']} 条。</p>"
        "<h2>平台对比</h2><table><tr><th>平台</th><th>样本</th><th>互动率</th></tr>"
        f"{platform_rows}</table><h2>时间组对比</h2><pre>{times}</pre>"
    )
    return StreamingResponse(
        iter([html.encode("utf-8")]),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=contentpilot-report.html"},
    )


@router.post("/analytics/ai-summary")
def ai_summary(
    request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)
) -> dict:
    summary = summarize(db)
    platforms = grouped(db, "platform")
    times = grouped(db, "group_type")
    if not summary["sampleCount"]:
        return success_response(
            request,
            {
                "summary": "暂无可分析数据",
                "keyFindings": [],
                "recommendations": ["先导入经过核验的平台互动数据"],
                "limitations": ["样本量为 0"],
                "provider": "RULE_BASED",
            },
        )
    best = platforms[0] if platforms else None
    return success_response(
        request,
        {
            "summary": (
                f"当前共 {summary['sampleCount']} 条样本，总互动率为 {summary['engagementRate']}%。"
            ),
            "keyFindings": [f"{best['name']} 当前样本互动率最高，为 {best['engagementRate']}%"]
            if best
            else [],
            "timeInsights": [f"{x['name']}：{x['engagementRate']}%" for x in times],
            "recommendations": [
                "持续补充真实数据后再判断时间推荐效果",
                "对高表现内容主题进行小规模复现",
            ],
            "limitations": (
                [f"包含 {summary['simulatedCount']} 条演示数据，不能作为真实效果结论"]
                if summary["simulatedCount"]
                else ["结论仅适用于当前已导入样本和统计口径"]
            ),
            "provider": "RULE_BASED",
        },
    )
